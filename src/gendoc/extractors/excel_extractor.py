"""Excel to Markdown extraction module.

Extracts product reference data from the Excel VBA file into structured Markdown files.
"""

import argparse
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# Excel sheet to output file mapping
SHEET_MAPPING = {
    "Paillasse": "paillasse.md",
    "Sorbonne": "sorbonne.md",
    "Revètement": "revetement.md",
    "Meubles": "meubles.md",
    "Tables EN": "tables-en.md",
    "Equipement": "equipement.md",
    "Elec sorb": "elec-sorb.md",
    "Compléments": "complements.md",
    "Fiches Existantes": "fiches-existantes.md",
}


def parse_image_position(value: str) -> Tuple[str, str, str, str]:
    """Parse image position string 'left,top,width,height' from row 2."""
    if not value or value.strip() == "":
        return ("", "", "", "")
    parts = [p.strip() for p in str(value).split(",")]
    if len(parts) == 4:
        return tuple(parts)
    return ("", "", "", "")


def clean_value(value) -> str:
    """Clean cell value - convert None to empty string."""
    if value is None:
        return ""
    return str(value).strip()


def extract_sheet_metadata(ws: Worksheet) -> Dict[str, Dict]:
    """Extract column metadata from rows 1-4.

    Returns dict: {col_letter: {type, prefix, shape_index, header}}
    """
    metadata = {}

    # Get max column from row 4 (headers)
    max_col = ws.max_column

    for col_idx in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)

        # Row 1: Type (TEXTE or IMAGE)
        col_type = clean_value(ws.cell(1, col_idx).value)

        # Row 2: Prefix for TEXTE or position for IMAGE
        row2_value = clean_value(ws.cell(2, col_idx).value)

        # Row 3: Shape index (should be numeric, but sometimes contains text labels)
        shape_index_raw = clean_value(ws.cell(3, col_idx).value)

        # Validate shape index - must be numeric or empty
        # Some columns (like "Famille de produit" in Meubles) have descriptive text here
        try:
            if shape_index_raw:
                int(shape_index_raw)  # Test if numeric
                shape_index = shape_index_raw
            else:
                shape_index = ""
        except ValueError:
            # Non-numeric value in row 3 - this is a label/description, not a shape index
            shape_index = ""

        # Row 4: Header
        header = clean_value(ws.cell(4, col_idx).value)

        metadata[col_letter] = {
            "type": col_type,
            "prefix": row2_value if col_type == "TEXTE" else "",
            "position": row2_value if col_type == "IMAGE" else "",
            "shape_index": shape_index,
            "header": header,
        }

    return metadata


def extract_product(ws: Worksheet, row_idx: int, metadata: Dict, family_name: str) -> Optional[Dict]:
    """Extract a single product from a row."""
    # Column A is always the code
    code = clean_value(ws.cell(row_idx, 1).value)

    # Skip empty rows
    if not code:
        return None

    product = {
        "code": code,
        "famille": family_name,
        "texte_fields": [],
        "dimension_fields": [],
        "image_fields": [],
        "metadata": [],
    }

    # Process each column based on metadata
    for col_idx in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        meta = metadata.get(col_letter, {})

        col_type = meta.get("type", "")
        header = meta.get("header", "")
        shape_index = meta.get("shape_index", "")
        value = clean_value(ws.cell(row_idx, col_idx).value)

        # Store in metadata
        product["metadata"].append({
            "field": header,
            "type": col_type,
            "prefix": meta.get("prefix", ""),
            "shape_index": shape_index,
        })

        # Store based on field name
        if col_idx == 1:
            product["code"] = value
        elif header.lower() in ["ref", "reference"]:
            product["ref"] = value
        elif header.lower() in ["titre", "title"]:
            product["titre"] = value
        elif col_type == "IMAGE":
            left, top, width, height = parse_image_position(meta.get("position", ""))
            product["image_fields"].append({
                "position": header,
                "path": value,
                "left": left,
                "top": top,
                "width": width,
                "height": height,
                "shape_index": shape_index,
            })
        elif col_type == "TEXTE":
            # Distinguish between main text and dimensions
            if header.lower() in ["texte", "texte complementaire", "mise_en_oeuvre", "finition", "commentaire"]:
                product["texte_fields"].append({
                    "name": header,
                    "value": value,
                })
            else:
                # Dimension or other field
                product["dimension_fields"].append({
                    "name": header,
                    "value": value,
                    "prefix": meta.get("prefix", ""),
                    "shape_index": shape_index,
                })

    # Set default values
    product.setdefault("ref", "")
    product.setdefault("titre", "")

    return product


def format_product_md(product: Dict, family_name: str) -> str:
    """Format a product as Markdown."""
    code = product["code"]

    # Special format for Fiches Existantes
    if family_name == "Fiches Existantes":
        md = f"## {code}\n\n"
        md += "| Champ | Valeur |\n"
        md += "|-------|--------|\n"
        md += f"| code | {product['code']} |\n"
        md += f"| ref | {product.get('ref', '')} |\n"
        md += f"| titre | {product.get('titre', '')} |\n"

        # Extract special fields from dimension_fields
        for field in product["dimension_fields"]:
            name = field["name"].lower()
            if name in ["repertoire", "fichier", "chemin", "nb pages"]:
                md += f"| {field['name']} | {field['value']} |\n"

        # Extract from texte_fields
        for field in product["texte_fields"]:
            name = field["name"].lower()
            if name == "commentaire":
                md += f"| {field['name']} | {field['value']} |\n"

        md += "\n---\n\n"
        return md

    # Standard format
    md = f"## {code}\n\n"
    md += "| Champ | Valeur |\n"
    md += "|-------|--------|\n"
    md += f"| code | {product['code']} |\n"
    md += f"| ref | {product.get('ref', '')} |\n"
    md += f"| titre | {product.get('titre', '')} |\n"
    md += f"| famille | {product['famille']} |\n\n"

    # Texte section
    md += "### Texte\n\n"
    if product["texte_fields"]:
        for field in product["texte_fields"]:
            if field["value"]:
                md += f"{field['value']}\n\n"
    else:
        md += "Aucune\n\n"

    # Dimensions section
    md += "### Dimensions\n\n"
    if product["dimension_fields"]:
        md += "| Dimension | Valeur | Prefix | Shape Index |\n"
        md += "|-----------|--------|--------|-------------|\n"
        for field in product["dimension_fields"]:
            md += f"| {field['name']} | {field['value']} | {field['prefix']} | {field['shape_index']} |\n"
        md += "\n"
    else:
        md += "Aucune\n\n"

    # Images section
    md += "### Images\n\n"
    if product["image_fields"]:
        md += "| Position | Chemin | Left | Top | Width | Height | Shape Index |\n"
        md += "|----------|--------|------|-----|-------|--------|-------------|\n"
        for field in product["image_fields"]:
            md += f"| {field['position']} | {field['path']} | {field['left']} | {field['top']} | {field['width']} | {field['height']} | {field['shape_index']} |\n"
        md += "\n"
    else:
        md += "Aucune\n\n"

    # Metadata PowerPoint section
    md += "### Metadata PowerPoint\n\n"
    md += "| Champ | Type | Prefix | Shape Index |\n"
    md += "|-------|------|--------|-------------|\n"
    for meta in product["metadata"]:
        md += f"| {meta['field']} | {meta['type']} | {meta['prefix']} | {meta['shape_index']} |\n"
    md += "\n---\n\n"

    return md


def extract_fiches_existantes(ws: Worksheet, family_name: str) -> Tuple[List[Dict], int]:
    """Extract Fiches Existantes - special case with different structure.

    This sheet doesn't have the 4-row metadata structure.
    Row 1 is headers, Row 2+ is data.
    """
    products = []

    # Get headers from row 1
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        header = clean_value(ws.cell(1, col_idx).value)
        headers[col_idx] = header

    # Extract products from row 2 onwards
    for row_idx in range(2, ws.max_row + 1):
        code = clean_value(ws.cell(row_idx, 1).value)
        if not code:
            continue

        product = {
            "code": code,
            "famille": family_name,
            "texte_fields": [],
            "dimension_fields": [],
            "image_fields": [],
            "metadata": [],
        }

        # Extract each field
        for col_idx, header in headers.items():
            value = clean_value(ws.cell(row_idx, col_idx).value)
            header_lower = header.lower()

            if header_lower == "code":
                product["code"] = value
            elif header_lower in ["référence", "reference"]:
                product["ref"] = value
            elif header_lower == "titre":
                product["titre"] = value
            elif header_lower in ["répertoire", "repertoire"]:
                product.setdefault("repertoire", value)
                product["dimension_fields"].append({
                    "name": "Répertoire",
                    "value": value,
                    "prefix": "",
                    "shape_index": "",
                })
            elif header_lower == "fichier":
                product.setdefault("fichier", value)
                product["dimension_fields"].append({
                    "name": "Fichier",
                    "value": value,
                    "prefix": "",
                    "shape_index": "",
                })
            elif header_lower == "chemin":
                product.setdefault("chemin", value)
                product["dimension_fields"].append({
                    "name": "Chemin",
                    "value": value,
                    "prefix": "",
                    "shape_index": "",
                })
            elif header_lower == "nb pages":
                product.setdefault("nb_pages", value)
                product["dimension_fields"].append({
                    "name": "Nb pages",
                    "value": value,
                    "prefix": "",
                    "shape_index": "",
                })
            elif header_lower == "commentaire":
                product["texte_fields"].append({
                    "name": "Commentaire",
                    "value": value,
                })

        # Ensure defaults
        product.setdefault("ref", "")
        product.setdefault("titre", "")

        products.append(product)

    # Sort by code
    products.sort(key=lambda p: p["code"])

    return products, len(products)


def extract_family(ws: Worksheet, family_name: str) -> Tuple[List[Dict], int]:
    """Extract all products from a family sheet.

    Returns (products, count).
    """
    # Special handling for Fiches Existantes
    if family_name == "Fiches Existantes":
        return extract_fiches_existantes(ws, family_name)

    metadata = extract_sheet_metadata(ws)
    products = []

    # Start from row 5 (data rows)
    for row_idx in range(5, ws.max_row + 1):
        product = extract_product(ws, row_idx, metadata, family_name)
        if product:
            products.append(product)

    # Sort by code
    products.sort(key=lambda p: p["code"])

    return products, len(products)


def write_family_md(output_dir: Path, family_name: str, sheet_name: str, products: List[Dict], extraction_date: str):
    """Write a family MD file."""
    filename = SHEET_MAPPING[sheet_name]
    filepath = output_dir / filename

    md = f"# {family_name}\n\n"
    md += f"> Extracted from: Génération Fiche Technique DELAGRAVE.xlsm\n"
    md += f"> Sheet: {sheet_name}\n"
    md += f"> Date: {extraction_date}\n"
    md += f"> Total references: {len(products)}\n\n"

    for product in products:
        md += format_product_md(product, family_name)

    filepath.write_text(md, encoding="utf-8")


def extract_parametrage(wb: openpyxl.Workbook, output_dir: Path, extraction_date: str):
    """Extract the Parametrage sheet."""
    if "Paramétrage" not in wb.sheetnames:
        print("Warning: Paramétrage sheet not found")
        return

    ws = wb["Paramétrage"]

    md = "# Parametrage\n\n"
    md += "> Configuration extraite du fichier Excel\n"
    md += "> Mapping famille -> template PowerPoint\n\n"
    md += "| Famille | Template | Layout | Type | Data Type |\n"
    md += "|---------|----------|--------|------|-----------|"

    # Find the Parametre_Onglet table (usually starts around row 2-5)
    # Look for headers
    for row_idx in range(1, 20):
        cell_value = clean_value(ws.cell(row_idx, 1).value)
        if "famille" in cell_value.lower() or "onglet" in cell_value.lower():
            # Found header row, extract data below
            header_row = row_idx
            for data_row in range(header_row + 1, ws.max_row + 1):
                famille = clean_value(ws.cell(data_row, 1).value)
                if not famille:
                    break
                template = clean_value(ws.cell(data_row, 2).value)
                layout = clean_value(ws.cell(data_row, 3).value)
                type_val = clean_value(ws.cell(data_row, 4).value)
                data_type = clean_value(ws.cell(data_row, 5).value) if ws.max_column >= 5 else ""

                md += f"\n| {famille} | {template} | {layout} | {type_val} | {data_type} |"
            break

    md += "\n\n## Chemins reseau (originaux)\n\n"
    md += "| Cle | Valeur |\n"
    md += "|-----|--------|\n"

    # Look for path configuration (usually labeled as REPERTOIRE, etc.)
    for row_idx in range(1, min(50, ws.max_row + 1)):
        key = clean_value(ws.cell(row_idx, 1).value)
        if key and ("REPERTOIRE" in key.upper() or "CHEMIN" in key.upper() or "PATH" in key.upper()):
            value = clean_value(ws.cell(row_idx, 2).value)
            md += f"| {key} | {value} |\n"

    filepath = output_dir / "_parametrage.md"
    filepath.write_text(md, encoding="utf-8")


def generate_index(output_dir: Path, family_counts: Dict[str, int], extraction_date: str):
    """Generate the master index file."""
    total_count = sum(family_counts.values())
    family_count = len(family_counts)

    md = "# Index des References Delagrave\n\n"
    md += f"> Source: Génération Fiche Technique DELAGRAVE.xlsm\n"
    md += f"> Extraction: {extraction_date}\n"
    md += f"> Total: {total_count} references dans {family_count} familles\n\n"
    md += "## Familles\n\n"
    md += "| Famille | Fichier | Nb references | Type |\n"
    md += "|---------|---------|---------------|------|\n"

    # Define type info
    type_info = {
        "Paillasse": "PPT (texte + dimensions + image)",
        "Sorbonne": "PPT (texte + dimensions + image)",
        "Revètement": "PPT (texte + 2 images)",
        "Meubles": "PPT (texte + image)",
        "Tables EN": "PPT (texte + image)",
        "Equipement": "PPT (images positionnees)",
        "Elec sorb": "PPT (images positionnees)",
        "Compléments": "PPT (images positionnees)",
        "Fiches Existantes": "EXT (fichiers .pptx)",
    }

    for sheet_name, filename in SHEET_MAPPING.items():
        count = family_counts.get(sheet_name, 0)
        family_type = type_info.get(sheet_name, "PPT")
        md += f"| {sheet_name} | [{filename}]({filename}) | {count} | {family_type} |\n"

    md += "\n## Structure d'un fichier famille\n\n"
    md += "Chaque fichier MD contient :\n"
    md += "- En-tête : nom famille, source, date, compteur\n"
    md += "- Sections `## {Code}` : une par produit\n"
    md += "  - Tableau identité : code, ref, titre, famille\n"
    md += "  - `### Texte` : contenu descriptif\n"
    md += "  - `### Dimensions` : tableau avec valeur, prefix PPTX, shape index\n"
    md += "  - `### Images` : tableau avec chemin, position (left/top/width/height), shape index\n"
    md += "  - `### Metadata PowerPoint` : mapping complet colonnes -> shapes\n\n"

    md += "## Fichiers speciaux\n\n"
    md += "| Fichier | Role |\n"
    md += "|---------|------|\n"
    md += "| [_index.md](_index.md) | Ce fichier - point d'entree |\n"
    md += "| [_parametrage.md](_parametrage.md) | Config mapping famille -> template PowerPoint |\n"

    filepath = output_dir / "_index.md"
    filepath.write_text(md, encoding="utf-8")


def extract_all(excel_path: Path, output_dir: Path, dry_run: bool = False, family_filter: Optional[str] = None) -> Dict[str, int]:
    """Extract all families from Excel to Markdown.

    Args:
        excel_path: Path to Excel file
        output_dir: Output directory for MD files
        dry_run: If True, only print what would be extracted
        family_filter: If set, only extract this specific family

    Returns:
        Dict of family name -> count
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    extraction_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    family_counts = {}

    # Create output directory
    if not dry_run:
        output_dir.mkdir(parents=True, exist_ok=True)

    # Extract each family
    for sheet_name, filename in SHEET_MAPPING.items():
        # Apply family filter
        if family_filter and sheet_name != family_filter:
            continue

        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in Excel file")
            continue

        print(f"Extracting family: {sheet_name}...")
        ws = wb[sheet_name]
        products, count = extract_family(ws, sheet_name)
        family_counts[sheet_name] = count

        if dry_run:
            print(f"  Would extract {count} products to {filename}")
        else:
            write_family_md(output_dir, sheet_name, sheet_name, products, extraction_date)
            print(f"  Extracted {count} products to {filename}")

    # Extract parametrage
    if not family_filter:
        if not dry_run:
            print("Extracting parametrage...")
            extract_parametrage(wb, output_dir, extraction_date)

        # Generate index
        if not dry_run:
            print("Generating index...")
            generate_index(output_dir, family_counts, extraction_date)

    total = sum(family_counts.values())
    print(f"\nTotal: {total} references across {len(family_counts)} families")

    return family_counts


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(description="Extract product references from Excel to Markdown")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be extracted without writing files")
    parser.add_argument("--family", type=str, help="Extract only a specific family")
    parser.add_argument("--excel", type=str, help="Path to Excel file (default: auto-detect)")
    parser.add_argument("--output", type=str, help="Output directory (default: Delagrave/references/)")

    args = parser.parse_args()

    # Determine paths
    if args.excel:
        excel_path = Path(args.excel)
    else:
        # Auto-detect Excel file
        base_dir = Path(__file__).parent.parent.parent.parent
        excel_path = base_dir / "Delagrave" / "Tables_de_references_produits" / "Génération Fiche Technique DELAGRAVE.xlsm"

    if args.output:
        output_dir = Path(args.output)
    else:
        base_dir = Path(__file__).parent.parent.parent.parent
        output_dir = base_dir / "Delagrave" / "references"

    try:
        extract_all(excel_path, output_dir, dry_run=args.dry_run, family_filter=args.family)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
